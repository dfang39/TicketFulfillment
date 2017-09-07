import datetime
import os
import openpyxl

today = datetime.date.today().strftime("%m/%d/%Y")

desktop_folder = 'C:\\Users\\DanFang\\Desktop'
os.chdir(desktop_folder)

requests_book = openpyxl.load_workbook('TicketRequests2.xlsx', read_only=True)
request_sheet = requests_book.get_sheet_by_name('Ticket Request Associated View')


class Request(object):
    def __init__(self, id_number="", agency="", tickets_requested="", request_status="", number_children="", number_adults="", rating="", number_of_assignments=0, ticket_assignments=[]):
        self.id_number = id_number
        self.agency = agency
        self.tickets_requested = tickets_requested
        self.request_status = request_status
        self.number_children = number_children
        self.number_adults = number_adults
        self.rating = rating
        self.number_of_assignments = number_of_assignments
        self.ticket_assignments = ticket_assignments


request_objects = []
for row in range(2, request_sheet.max_row + 1):
    rating = 0
    if request_sheet['O' + str(row)].value == 'Silver':
        rating = 1
    elif request_sheet['O' + str(row)].value == 'Bronze':
        rating = 2
    request_objects.append(
        Request(request_sheet['D' + str(row)].value, request_sheet['E' + str(row)].value, request_sheet['J' + str(row)].value, request_sheet['G' + str(row)].value, request_sheet['N' + str(row)].value, request_sheet['P' + str(row)].value, rating))

request_objects.sort(key=lambda r: (r.rating, r.tickets_requested))
print("ordering by agency ranking, then by size from smallest to largest")

donations_book = openpyxl.load_workbook('TicketDonations1.xlsx', read_only=True)
donation_sheet = donations_book.get_sheet_by_name('Ticket Donation Associated View')


class Donation(object):
    def __init__(self, id_number="", contact="", tickets_donated="", ticket_assignments=[]):
        self.id_number = id_number
        self.contact = contact
        self.tickets_donated = tickets_donated
        self.ticket_assignments = ticket_assignments


donation_objects = []
for row in range(2, donation_sheet.max_row + 1):
    donation_objects.append(
        Donation(donation_sheet['D' + str(row)].value, donation_sheet['L' + str(row)].value, donation_sheet['G' + str(row)].value))



donation_objects.sort(key=lambda r: r.tickets_donated)


print([request.tickets_requested for request in request_objects])
print(sum([request.tickets_requested for request in request_objects]))
print([donation.tickets_donated for donation in donation_objects])
print(sum([donation.tickets_donated for donation in donation_objects]))

class Assignment(object):
    def __init__(self, request_id="", donation_id="", quantity=""):
        self.request_id = request_id
        self.donation_id = donation_id
        self.quantity = quantity


def match_exact(request):
    print('beginning exact match')
    donations = [donation for donation in donation_objects if donation.tickets_donated == request.tickets_requested]
    if len(donations) == 0:
        print("no exact match")
        return request.tickets_requested
    else:
        assigned_tickets = request.tickets_requested
        request.ticket_assignments.append(Assignment(request.id_number, donations[0].id_number, assigned_tickets))
        request.number_of_assignments += 1
        donations[0].ticket_assignments.append(Assignment(request.id_number, donations[0].id_number, assigned_tickets))
        print(request.id_number, 'had', request.tickets_requested, 'in the request and was assigned', assigned_tickets,
                  'tickets from', donations[0].id_number, 'matching exactly')
        request.tickets_requested -= assigned_tickets
        donations[0].tickets_donated -= assigned_tickets
        print("exact match found")
        return request.tickets_requested


def match_fuzzy(request):
    tickets_remaining = match_exact(request)
    print('tried to exact match, now beginning fuzzy match')
    if tickets_remaining == 0:
        print("finished matching with exact matches")
        return "finished matching with exact matches"

    adults = request.number_adults

    smaller_donations = [donation for donation in donation_objects if
                         request.tickets_requested >= donation.tickets_donated > 1]
    smaller_donations.sort(key=lambda r: r.tickets_donated, reverse=True)
    print('smaller donations', len(smaller_donations))
    larger_donations = [donation for donation in donation_objects if
                        donation.tickets_donated >= request.tickets_requested]
    print('larger donations', len(larger_donations))
    if adults == 1 or len(smaller_donations) == 0:
        print("only one chaperone or no more smaller donations")
        if len(larger_donations) == 0:
            print("no more tickets")
            return "no more tickets"
        else:
            print('beginning larger donations')
            for donation in larger_donations:
                donated_tickets = donation.tickets_donated
                if 0 < request.tickets_requested <= donated_tickets:
                    assigned_tickets = request.tickets_requested
                    request.ticket_assignments.append(
                        Assignment(request.id_number, donation.id_number, assigned_tickets))
                    request.number_of_assignments += 1
                    donation.ticket_assignments.append(
                        Assignment(request.id_number, donation.id_number, assigned_tickets))
                    donation.tickets_donated -= assigned_tickets
                    donated_tickets -= assigned_tickets
                    print(request.id_number, 'had', request.tickets_requested, 'in the request and was assigned',
                          assigned_tickets,
                          'tickets from', donation.id_number, ', fulfilling the request and leaving', donated_tickets,
                          'in the donation')
                    request.tickets_requested -= assigned_tickets
                    if request.tickets_requested == 0:
                        return "All matched"
    elif adults > 1:
        for donation in smaller_donations:
            donated_tickets = donation.tickets_donated
            if request.tickets_requested >= donated_tickets > 1 and (request.tickets_requested - donated_tickets) > 1:
                assigned_tickets = donated_tickets
                adults = request.number_adults
                request.ticket_assignments.append(Assignment(request.id_number, donation.id_number, assigned_tickets))
                request.number_of_assignments += 1
                donation.ticket_assignments.append(Assignment(request.id_number, donation.id_number, assigned_tickets))
                adults -= 1
                request.number_adults -= 1

                print(request.id_number, 'had', request.tickets_requested, 'in the request and was assigned',
                              assigned_tickets, 'tickets from', donation.id_number,
                              ', taking all of the donation and leaving', adults, 'chaperones and', request.tickets_requested - assigned_tickets,
                              'tickets on the request')
                request.tickets_requested -= assigned_tickets
                donation.tickets_donated -= assigned_tickets
                donated_tickets -= assigned_tickets
                tickets_remaining -= assigned_tickets
                if tickets_remaining == 0:
                    print("request fulfilled")
                    return "tickets assigned"
                else:
                    match_fuzzy(request)


for request in request_objects:
    print(request.id_number, 'has', request.tickets_requested, 'in the request and', request.number_adults, 'adults')
    match_fuzzy(request)

    print('---')



print([request.tickets_requested for request in request_objects])
print(sum([request.tickets_requested for request in request_objects]))
print([donation.tickets_donated for donation in donation_objects])
print(sum([donation.tickets_donated for donation in donation_objects]))


print('fully assigned', len([request for request in request_objects if request.tickets_requested == 0]))
print('fully denied', len([request for request in request_objects if request.number_of_assignments == 0]))
print('approved with 1 set of tix', len([request for request in request_objects if request.number_of_assignments == 1]))
print('approved with 2 set of tix', len([request for request in request_objects if request.number_of_assignments == 2]))
print('approved with 3 set of tix', len([request for request in request_objects if request.number_of_assignments == 3]))
print('approved with 4+ set of tix', len([request for request in request_objects if request.number_of_assignments >= 4]))
