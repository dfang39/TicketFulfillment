import datetime
import os
import openpyxl

today = datetime.date.today().strftime("%m/%d/%Y")

desktop_folder = 'C:\\Users\\DanFang\\Desktop'
os.chdir(desktop_folder)

requests_book = openpyxl.load_workbook('TicketRequests2.xlsx', read_only=True)
request_sheet = requests_book.get_sheet_by_name('Ticket Request Associated View')


class Request(object):
    def __init__(self, id_number="", agency="", tickets_requested="", request_status="", number_children="", number_adults="", rating="", ticket_assignments=[]):
        self.id_number = id_number
        self.agency = agency
        self.tickets_requested = tickets_requested
        self.request_status = request_status
        self.number_children = number_children
        self.number_adults = number_adults
        self.rating = rating
        self.ticket_assignments = ticket_assignments


request_objects = []
for row in range(2, request_sheet.max_row + 1):
    rating = 0
    if request_sheet['O' + str(row)].value == 'Silver':
        rating = 1
    elif request_sheet['O' + str(row)].value == 'Bronze':
        rating = 2
    request_objects.append(
        Request(request_sheet['D' + str(row)].value, request_sheet['E' + str(row)].value, request_sheet['J' + str(row)].value, request_sheet['G' + str(row)].value, request_sheet['N' + str(row)].value, request_sheet['P' + str(row)].value, rating))

request_objects.sort(key=lambda r: r.tickets_requested)

request_objects_2 = []
request_objects_4 = []

for request in request_objects:
    if request.tickets_requested == 2 and request.number_children == request.number_adults:
        request_objects_2.append(request)
    elif request.tickets_requested == 4 and request.number_children == request.number_adults:
        request_objects_4.append(request)


donations_book = openpyxl.load_workbook('TicketDonations1.xlsx', read_only=True)
donation_sheet = donations_book.get_sheet_by_name('Ticket Donation Associated View')


class Donation(object):
    def __init__(self, id_number="", contact="", tickets_donated="", ticket_assignments=[]):
        self.id_number = id_number
        self.contact = contact
        self.tickets_donated = tickets_donated
        self.ticket_assignments = ticket_assignments


donation_objects = []
for row in range(2, donation_sheet.max_row + 1):
    donation_objects.append(
        Donation(donation_sheet['D' + str(row)].value, donation_sheet['L' + str(row)].value, donation_sheet['G' + str(row)].value))


donation_objects_2 = []
donation_objects_4 = []

for donation in donation_objects:
    if donation.tickets_donated == 2:
        donation_objects_2.append(donation)
    elif donation.tickets_donated == 4:
        donation_objects_4.append(donation)

donation_objects.sort(key=lambda r: r.tickets_donated)

request_2_count = len(request_objects_2)
request_4_count = len(request_objects_4)
donation_2_count = len(donation_objects_2)
donation_4_count = len(donation_objects_4)

print(request_2_count)
print(request_4_count)
print(donation_2_count)
print(donation_4_count)

print([request.tickets_requested for request in request_objects])
print([donation.tickets_donated for donation in donation_objects])


class Assignment(object):
    def __init__(self, request_id="", donation_id="", quantity=""):
        self.request_id = request_id
        self.donation_id = donation_id
        self.quantity = quantity


for request in request_objects_2:
    tickets = request.tickets_requested
    if tickets > 0:
        for donation in donation_objects_2:
            donations = donation.tickets_donated
            if donations > 0 and tickets > 0:
                assignments = donations
                request.ticket_assignments.append(Assignment(request.id_number, donation.id_number, assignments))
                donation.ticket_assignments.append(Assignment(request.id_number, donation.id_number, assignments))
                request.tickets_requested -= assignments
                tickets -= assignments
                donation.tickets_donated -= assignments
                donations -= assignments
                print(request.id_number, 'was assigned 2 tickets from', donation.id_number, 'matching exactly')
    print(" --- ")

for request in request_objects_4:
    tickets = request.tickets_requested
    if tickets > 0:
        for donation in donation_objects_4:
            assignments = donation.tickets_donated
            donations = donation.tickets_donated
            if donations > 0 and tickets > 0:
                request.ticket_assignments.append(Assignment(request.id_number, donation.id_number, assignments))
                donation.ticket_assignments.append(Assignment(request.id_number, donation.id_number, assignments))
                request.tickets_requested -= assignments
                tickets -= assignments
                donation.tickets_donated -= assignments
                donations -= assignments
                print(request.id_number, 'was assigned 4 tickets from', donation.id_number, 'matching exactly')
        for donation in donation_objects_2:
            assignments = donation.tickets_donated
            donations = donation.tickets_donated
            if donations > 0 and tickets > 0:
                request.ticket_assignments.append(Assignment(request.id_number, donation.id_number, assignments))
                donation.ticket_assignments.append(Assignment(request.id_number, donation.id_number, assignments))
                request.tickets_requested -= assignments
                tickets -= assignments
                donation.tickets_donated -= assignments
                donations -= assignments
                print(request.id_number, 'was assigned 2 tickets from', donation.id_number, 'and has', tickets, 'left on the request')
    print(" --- ")


request_objects.sort(key=lambda r: r.rating)
print("reordering by agency ranking")
donation_objects.sort(key=lambda r: r.tickets_donated, reverse=True)
print("reordered ticket donation from largest to smallest")


for request in request_objects:
    requested_tickets = request.tickets_requested
    if requested_tickets > 0:
        for donation in donation_objects:
            donated_tickets = donation.tickets_donated
            if requested_tickets > donated_tickets > 1 and (requested_tickets - donated_tickets) > 1:
                assigned_tickets = donated_tickets
                adults = request.number_adults
                print(request.id_number, 'initially requested', requested_tickets, 'and had', adults, 'chaperones')
                if adults > 1:
                    request.ticket_assignments.append(Assignment(request.id_number, donation.id_number, assigned_tickets))
                    donation.ticket_assignments.append(Assignment(request.id_number, donation.id_number, assigned_tickets))
                    adults -= 1
                    request.number_adults -= 1
                    requested_tickets -= assigned_tickets
                    print(request.id_number, 'had', request.tickets_requested, 'in the request and was assigned',
                          assigned_tickets, 'tickets from', donation.id_number,
                          'taking all of the donation and leaving', adults, 'chaperones and', requested_tickets,
                          'tickets left on the request')
                    request.tickets_requested -= assigned_tickets
                    donation.tickets_donated -= assigned_tickets
                    donated_tickets -= assigned_tickets
                elif adults == 1:
                    print('finding a matching donation...')
                    continue
                if requested_tickets == 0:
                    break
            elif requested_tickets <= donated_tickets:
                assigned_tickets = requested_tickets
                request.ticket_assignments.append(Assignment(request.id_number, donation.id_number, assigned_tickets))
                donation.ticket_assignments.append(Assignment(request.id_number, donation.id_number, assigned_tickets))
                donation.tickets_donated -= assigned_tickets
                donated_tickets -= assigned_tickets
                print(request.id_number, 'had', requested_tickets, 'in the request and was assigned', assigned_tickets,
                      'tickets from', donation.id_number, 'fulfilling the request and leaving', donated_tickets, 'in the donation')
                request.tickets_requested -= assigned_tickets
                requested_tickets -= assigned_tickets
                if requested_tickets == 0:
                    break
    print('---')

print([request.tickets_requested for request in request_objects])
print([donation.tickets_donated for donation in donation_objects])

# for request in request_objects:
    # print(len(request.ticket_assignments))