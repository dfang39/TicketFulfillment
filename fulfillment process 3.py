import datetime
import os
import openpyxl

today = datetime.date.today().strftime("%m/%d/%Y")

desktop_folder = 'C:\\Users\\DanFang\\Desktop'
os.chdir(desktop_folder)

requests_book = openpyxl.load_workbook('TicketRequests2.xlsx', read_only=True)
request_sheet = requests_book.get_sheet_by_name('Ticket Request Associated View')


class Request(object):
    def __init__(self, id_number="", agency="", tickets_requested="", request_status="", number_children="", number_adults="", rating="", number_of_assignments=0, ticket_assignments=[]):
        self.id_number = id_number
        self.agency = agency
        self.tickets_requested = tickets_requested
        self.request_status = request_status
        self.number_children = number_children
        self.number_adults = number_adults
        self.rating = rating
        self.number_of_assignments = number_of_assignments
        self.ticket_assignments = ticket_assignments


request_objects = []
for row in range(2, request_sheet.max_row + 1):
    rating = 0
    if request_sheet['O' + str(row)].value == 'Silver':
        rating = 1
    elif request_sheet['O' + str(row)].value == 'Bronze':
        rating = 2
    request_objects.append(
        Request(request_sheet['D' + str(row)].value, request_sheet['E' + str(row)].value, request_sheet['J' + str(row)].value, request_sheet['G' + str(row)].value, request_sheet['N' + str(row)].value, request_sheet['P' + str(row)].value, rating))

request_objects.sort(key=lambda r: (r.rating, r.tickets_requested))
print("ordering by agency ranking, then by size from smallest to largest")

donations_book = openpyxl.load_workbook('TicketDonations1.xlsx', read_only=True)
donation_sheet = donations_book.get_sheet_by_name('Ticket Donation Associated View')


class Donation(object):
    def __init__(self, id_number="", contact="", tickets_donated="", ticket_assignments=[]):
        self.id_number = id_number
        self.contact = contact
        self.tickets_donated = tickets_donated
        self.ticket_assignments = ticket_assignments


donation_objects = []
for row in range(2, donation_sheet.max_row + 1):
    donation_objects.append(
        Donation(donation_sheet['D' + str(row)].value, donation_sheet['L' + str(row)].value, donation_sheet['G' + str(row)].value))



donation_objects.sort(key=lambda r: r.tickets_donated)


print([request.tickets_requested for request in request_objects])
print(sum([request.tickets_requested for request in request_objects]))
print([donation.tickets_donated for donation in donation_objects])
print(sum([donation.tickets_donated for donation in donation_objects]))

class Assignment(object):
    def __init__(self, request_id="", donation_id="", quantity=""):
        self.request_id = request_id
        self.donation_id = donation_id
        self.quantity = quantity


for request in request_objects:
    requested_tickets = request.tickets_requested
    if requested_tickets > 0:
        for donation in donation_objects:
            donated_tickets = donation.tickets_donated
            if donated_tickets == requested_tickets:
                assigned_tickets = requested_tickets
                request.ticket_assignments.append(Assignment(request.id_number, donation.id_number, assigned_tickets))
                request.number_of_assignments += 1
                donation.ticket_assignments.append(Assignment(request.id_number, donation.id_number, assigned_tickets))
                print(request.id_number, 'had', requested_tickets, 'in the request and was assigned', assigned_tickets,
                      'tickets from', donation.id_number, 'matching exactly')
                request.tickets_requested -= assigned_tickets
                requested_tickets -= assigned_tickets
                donation.tickets_donated -= assigned_tickets
                donated_tickets -= assigned_tickets
                break


    print("---")


donation_objects.sort(key=lambda r: r.tickets_donated, reverse=True)
print("reordered ticket donation from largest to smallest")

print([request.tickets_requested for request in request_objects])
print([donation.tickets_donated for donation in donation_objects])

for request in request_objects:
    requested_tickets = request.tickets_requested
    adults = request.number_adults
    if requested_tickets > 0:
        print(request.id_number, 'initially requested', requested_tickets, 'and had', adults, 'chaperones')
        donation_smaller = [donation for donation in donation_objects if
                            requested_tickets >= donation.tickets_donated > 1]
        for donation in donation_smaller:
            donated_tickets = donation.tickets_donated
            if requested_tickets >= donated_tickets and (requested_tickets - donated_tickets) > 1:
                assigned_tickets = donated_tickets
                adults = request.number_adults
                if adults > 1:
                    request.ticket_assignments.append(Assignment(request.id_number, donation.id_number, assigned_tickets))
                    request.number_of_assignments += 1
                    donation.ticket_assignments.append(Assignment(request.id_number, donation.id_number, assigned_tickets))
                    adults -= 1
                    request.number_adults -= 1
                    requested_tickets -= assigned_tickets
                    print(request.id_number, 'had', request.tickets_requested, 'in the request and was assigned',
                          assigned_tickets, 'tickets from', donation.id_number,
                          ', taking all of the donation and leaving', adults, 'chaperones and', requested_tickets,
                          'tickets on the request')
                    request.tickets_requested -= assigned_tickets
                    donation.tickets_donated -= assigned_tickets
                    donated_tickets -= assigned_tickets
                elif adults == 1:
                    print('finding a matching donation...')
                    continue
                if requested_tickets == 0:
                    break
            elif donated_tickets == requested_tickets:
                assigned_tickets = requested_tickets
                request.ticket_assignments.append(Assignment(request.id_number, donation.id_number, assigned_tickets))
                request.number_of_assignments += 1
                donation.ticket_assignments.append(Assignment(request.id_number, donation.id_number, assigned_tickets))
                print(request.id_number, 'had', requested_tickets, 'in the request and was assigned', assigned_tickets,
                      'tickets from', donation.id_number, 'matching exactly')
                request.tickets_requested -= assigned_tickets
                requested_tickets -= assigned_tickets
                donation.tickets_donated -= assigned_tickets
                donated_tickets -= assigned_tickets

        donation_larger = [donation for donation in donation_objects if
                            requested_tickets < donation.tickets_donated]
        donation_larger.sort(key=lambda r: r.tickets_donated)
        for donation in donation_larger:
            donated_tickets = donation.tickets_donated
            if 0 < requested_tickets <= donated_tickets:
                assigned_tickets = requested_tickets
                request.ticket_assignments.append(Assignment(request.id_number, donation.id_number, assigned_tickets))
                request.number_of_assignments += 1
                donation.ticket_assignments.append(Assignment(request.id_number, donation.id_number, assigned_tickets))
                donation.tickets_donated -= assigned_tickets
                donated_tickets -= assigned_tickets
                print(request.id_number, 'had', requested_tickets, 'in the request and was assigned', assigned_tickets,
                      'tickets from', donation.id_number, ', fulfilling the request and leaving', donated_tickets, 'in the donation')
                request.tickets_requested -= assigned_tickets
                requested_tickets -= assigned_tickets
                if requested_tickets == 0:
                    break

    print('---')



print([request.tickets_requested for request in request_objects])
print(sum([request.tickets_requested for request in request_objects]))
print([donation.tickets_donated for donation in donation_objects])
print(sum([donation.tickets_donated for donation in donation_objects]))


print(len([request for request in request_objects if request.tickets_requested == 0]))
print(len([request for request in request_objects if request.number_of_assignments == 0]))
print(len([request for request in request_objects if request.number_of_assignments == 1]))
print(len([request for request in request_objects if request.number_of_assignments == 2]))
print(len([request for request in request_objects if request.number_of_assignments == 3]))
print(len([request for request in request_objects if request.number_of_assignments >= 4]))
